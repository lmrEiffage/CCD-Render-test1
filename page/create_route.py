#Lucas Martínez Rodríguez - Eiffage

#IMPORTANTE: 
#DAR PERMISOS AL DIRECTORIO DE LECTURA, ESCRITURA Y EJECUCIÓN !!!
#CUIDADO AL PARSEAR CON LOS ESPACIOS DE NOMBRE DEL ARCHIVO ¡¡¡
#CUIDADO CON LOS ARCHIVOS QUE SE PARSEAN (XML.etree no tiene ningún tipo de forma de detectar intenciones maliciosas dentre de un kml)
#Al INICIAR EL PROGRAMA TO DO LO QUE NO SEA UNA CARPETA DEL MISMO SE BORRA DEL DIRECTORIO¡¡¡

import folium.map
import folium, os, time, shutil, simplekml, io
import pandas as pd
import xml.etree.ElementTree as ET
from streamlit_folium import st_folium
from streamlit_folium import st_folium
from openpyxl import load_workbook
from st_aggrid.grid_options_builder import GridOptionsBuilder
from st_aggrid import AgGrid, GridUpdateMode,ColumnsAutoSizeMode, GridOptionsBuilder, JsCode
import streamlit as st
from folium.plugins import Fullscreen, Draw
import streamlit.components.v1 as components
import hydralit_components as hc
from folium.map import Marker
from pathlib import Path
import zipfile
from tqdm import tqdm  
import requests
import numpy as np

#Directorio del workspace 
dirname = os.path.dirname(__file__)

#Out of pages
parent_dir = os.path.dirname(dirname)

#landmarks
land_loc = os.path.join(parent_dir, "landmarks.csv")

def loading():
    progress_bar = st.progress(0)
    with st.spinner('Loading...'):
        time.sleep(1)
    progress_bar.progress(100)
    progress_bar.empty()
    
@st.cache_data
def prepare_data(): 
    loading()
    columns = ['name', 'longitude', 'latitude', 'altitud']
    
    path_csv = os.path.join(parent_dir, 'landmarks.csv')
    if os.path.exists(path_csv):
        empty_df = pd.DataFrame(columns=columns)
        empty_df.to_csv(path_csv, index=False) 
    else: 
        df = pd.DataFrame(columns=columns)
        df.to_csv(path_csv, index=False)
        
    if 'mapa' not in st.session_state:
       st.session_state.map_config = {"center": [40.4637, 3.7492], "zoom": 5.5}
    
    if "markers_set" not in st.session_state:
        st.session_state["markers_set"] = set()
    
    if "manual_set" not in st.session_state:
        st.session_state["manual_set"] = []
        
    if 'new' not in st.session_state:
        st.session_state['new'] = False

    if 'route_name' not in st.session_state:
        st.session_state['route_name'] = None
        
    if 'zoom' not in st.session_state:
        st.session_state['zoom'] = None
        
    if 'name' not in st.session_state:
        st.session_state['name'] = None
        
    if 'waypoints' not in st.session_state:
        st.session_state['waypoints'] = []
        
    if 'key' not in st.session_state:
        st.session_state['key'] = 0
        
    if 'mapa2' not in st.session_state:
        st.session_state.map_config = {"center": [40.4637, 3.7492], "zoom": 5.5}
        
    if 'changes_made' not in st.session_state:
        st.session_state['changes_made'] = 0
        
    if 'modo' not in st.session_state:
        st.session_state['modo'] = []
        
    if 'last_clicked' not in st.session_state:
        st.session_state['last_clicked'] = None
        
    if 'output' not in st.session_state:
        st.session_state['output'] = None
        
    if "modo_man" not in st.session_state:
        st.session_state["modo_man"] = 0
        
    if "df_manual" not in st.session_state: 
        st.session_state["df_manual"] = []
        
    if "plantas" not in st.session_state: 
        path_plantas = os.path.join(parent_dir, "data", "plantas.csv")
        plantas_df = pd.read_csv(path_plantas)
        st.session_state["plantas"] = plantas_df
        
    if "data_map" not in st.session_state:
        st.session_state["data_map"] = [40,3]

@st.cache_data
def prepare_data_route(route_name):
    if not route_name: 
        st.sidebar.error("ERROR: introduce a route name")
        return
    else:
        path = os.path.join(parent_dir, "Trayectorias", route_name)
        try:
            if not os.path.exists(path):
                os.makedirs(path)
                #df_vacio = pd.DataFrame()
                #path_csv = os.path.join(path, 'waypoints.csv')
                #df_vacio.to_csv(path_csv, index=False)
                create_excell(route_name, path)
                create_kml(route_name)

        except Exception as e: 
            return

def create_excell(route_name, path):
    plantilla = os.path.join(parent_dir, "data", 'plantilla.xlsx')
    wb = load_workbook(plantilla)
    ws = wb.active

    ws['I2'] = route_name

    destino = os.path.join(path, route_name + '.xlsx')
    wb.save(destino)

def load_kmz(file):
    with zipfile.ZipFile(file, "r") as kmz: 
        kmz_files = kmz.namelist()
        for file_name in kmz_files:
            if file_name.endswith('.kml'):
                kml_data = kmz.read(file_name)
                kml_content = io.BytesIO(kml_data)
                load_kml(kml_content)

def charge_file(files):
    if not isinstance(files, list):
        files = [files]

    for file in files:
        if file is not None:
            filename = file.name
            if 'kml' in filename:
                load_kml(file)
            elif 'csv' in filename:
                load_csv(file)
            elif 'kmz' in filename:
                load_kmz(file)
            else:
                st.sidebar.error("ERROR: invalid file")
        else:
            st.sidebar.error("ERROR: file is None")

def load_kml(file):
    waypoints = []
    try:
        tree = ET.parse(file)
        root = tree.getroot()

    except ET.ParseError as e:
        print(f"ERROR: parsing {e}")
        return

    namespace = {'kml': 'http://www.opengis.net/kml/2.2'}
    placemarks = root.findall('.//kml:Placemark', namespace)

    for placemark in placemarks:
        if placemark.find('kml:LineString', namespace) is None:
            name = placemark.find('kml:name', namespace)
            if name is not None:
                name = name.text.strip()
            coordinates_str = placemark.find('.//kml:coordinates', namespace).text.strip()
            coordinates_parts = coordinates_str.split(',')
        
            waypoint = {
            'name': name if name else None,
            'longitude': float(coordinates_parts[0]),
            'latitude': float(coordinates_parts[1]),
            'altitud':float(coordinates_parts[2])
            }
            st.session_state["markers_set"].add(create_marker(waypoint))
            waypoints.append(waypoint)
    add_to_csv(waypoints)

def add_to_csv(data):
    csvLocation = os.path.join(parent_dir, "landmarks.csv")
    try:
        df = pd.DataFrame(data)
        centinela = is_file_empty(csvLocation)
        
        if st.session_state["markers_set"] and centinela == False:
            df_existing = pd.read_csv(csvLocation,  encoding='utf-8')
            df_filtered = df[~((df['latitude'].isin(df_existing['latitude'])) & 
                   (df['longitude'].isin(df_existing['longitude'])))]
            df_vacio = pd.DataFrame(columns=['name', 'longitude', 'latitude', 'altitud', 'ASL'])
            df_vacio.to_csv(csvLocation, index=False)
            if not df_existing.empty:
                df_combined = pd.concat([df_existing, df_filtered], ignore_index=True)
                #df_combined['name'] = range(len(df_combined))
                df_combined['name'] = ['waypoint ' + str(i) for i in range(len(df_combined))]
                df_combined.to_csv(csvLocation, mode='a', header=False, index=False)
            else:
                df_filtered.to_csv(csvLocation, mode='a', header=False, index=False)
                
            actual = pd.read_csv(csvLocation, encoding='utf-8')
            df = ASL(actual)
            df['altitud'] = df.apply(lambda row: row['altitud'] - row['ASL'] if row['altitud'] > row['ASL'] else row['altitud'], axis=1)
            #print(df)
            df.to_csv(csvLocation, index=False)
            
        else:
            df.to_csv(csvLocation, index=False,  encoding='utf-8')

    except IOError as e:
        print(f"ERROR: Unable to open or write to CSV file '{csvLocation}': {e}")

    except Exception as e:
        print(f"Unexpected ERROR: {e}")

def is_file_empty(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        content_stripped = content.strip()
        if not content_stripped:
            return True
        return False

def create_marker(data, color = "blue"):
    return folium.Marker(location=[data["latitude"], data["longitude"]], popup=data["name"], tooltip=data["altitud"], icon=folium.Icon(color=color))

def load_csv(file):
    try:
        df = pd.read_csv(file)
        landmarks = []
        for index, row in df.iterrows():  
            landmark = {
                'name': row.get('name', 'Unknown'),
                'longitude': row.get('longitude', 0.0),
                'latitude': row.get('latitude', 0.0),
                'altitud':row.get('altitud', 0.0)
            }
            st.session_state["markers_set"].add(create_marker(landmark))
            landmarks.append(landmark)
        add_to_csv(landmarks)

    except Exception as e:  
        st.sidebar.error(f'Error loading file: {e}')

@st.cache_resource
def extract_landmarks(route_name):
    try:
        land_loc = os.path.join(parent_dir, route_name, "landmarks.csv")
        landmarks_df = pd.read_csv(land_loc, encoding='utf-8')
        return landmarks_df
    except:
        return []

@st.cache_data
def folium_map(landmarks, zoom = None, line = False):
    lat = 40.463667
    lon = -0.74922
    
    mapa = folium.Map(location=[lat, lon], zoom_start=5.5, zoom_control=False, scrollWheelZoom=False, dragging=False)

    if not landmarks.empty:
        if zoom is not None and len(landmarks)==1:
            mapa = folium.Map(location=[landmarks['latitude'].iloc[0], landmarks['longitude'].iloc[0]], zoom_start=zoom, zoom_control=True, scrollWheelZoom=True, dragging=True)

        iteraciones = 0
        points = []
        for _, landmark in landmarks.iterrows():
            points.append([landmark['latitude'], landmark['longitude']])
            if iteraciones == 0 and line == True:         
                folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(icon = 'star',color='red')).add_to(mapa) 
            
            else:
                if iteraciones == len(landmarks)-1 and line == True:
                    folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(icon = 'star',color='green')).add_to(mapa) 

                else: 
                    folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(color='blue')).add_to(mapa) 

            iteraciones += 1
            
        if line == True: 
            mapa.options.update({
                'zoom_control': True,
                'scrollWheelZoom': True,
                'dragging': True
            })
            folium.PolyLine(points,color="cyan",weight=10,opacity=1).add_to(mapa)    
            mapa.fit_bounds(mapa.get_bounds())

    return mapa

def centrar_en_landmark(landmarks_df, nombre_landmark):
    # Buscar el landmark por su nombre
    landmark = landmarks_df[landmarks_df['name'] == nombre_landmark].iloc[0]
    
    if  not landmark.empty:
        return [[landmark['name'], landmark['latitude'], landmark['longitude'], landmark['altitud']]]    
    else:
        st.warning(f"No se encontró el landmark '{nombre_landmark}'.")
        return []

def create_kml(route_name):
    filepath = os.path.join(parent_dir, "waypoints.csv")
    waypoints = pd.read_csv(filepath)
    kml = simplekml.Kml()
    kml_file = os.path.join(parent_dir, "Trayectorias", route_name, "{0}.kml".format(route_name))
    kmz_file = os.path.join(parent_dir, "Trayectorias", route_name, "{0}.kmz".format(route_name))
    
    for index, row in waypoints.iterrows():
        name = str(row['name'])
        lon = float(row['longitude'])
        lat = float(row['latitude'])
        alt = float(row['altitud'])
        asl = float(row['ASL'])
        
        # Crear un punto en el KML
        newpoint = kml.newpoint(name=name, coords=[(lon, lat, (alt+asl))])
        newpoint.altitudemode = simplekml.AltitudeMode.absolute    
    
    coords = [(float(row['longitude']), float(row['latitude']), (float(row['altitud']) + float(row['ASL']))) for index, row in waypoints.iterrows()]
    kml.newlinestring(name=route_name, coords=coords, altitudemode=simplekml.AltitudeMode.absolute)
    
    kml.save(kml_file)  
    
    with zipfile.ZipFile(kmz_file, 'w', zipfile.ZIP_DEFLATED) as kmz:
        kmz.write(kml_file, os.path.basename(kml_file))

def devolver_lat_lon(waypoints, num):
    if num == 0:
        latitude = waypoints.iloc[0, 1]  
        longitude = waypoints.iloc[0, 2]  
        
        latitude_str = str(latitude)
        longitude_str = str(longitude)
    
    elif num == 1: 
        last_row_data = waypoints.iloc[-1]
        latitude = last_row_data.iloc[1]  
        longitude = last_row_data.iloc[2]
        latitude_str = str(latitude)
        longitude_str = str(longitude)

    else:
        st.sidebar.error("ERROR") 
        print("Error")
        return

    return [latitude_str, longitude_str]

@st.cache_data
def lat_lon(zoom):
    df = pd.read_csv(land_loc)
    df_filtered = df[df['name'] == zoom]
    lat = df_filtered.iloc[0]["latitude"]
    lon = df_filtered.iloc[0]["longitude"]
    name = df_filtered.iloc[0]["name"]
    alt = df_filtered.iloc[0]["altitud"]
    
    return lat, lon, name, alt
    
@st.cache_data
def base_map(coord=[40,3], zoom = 5):
    m = folium.Map(location=coord, tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', attr='Esri', zoom_start = zoom, max_zoom = 20)
    Fullscreen().add_to(m)

    return m

@st.fragment(run_every=1)
def draw_map():
    m = base_map()
    fg = folium.FeatureGroup(name="Markers")
    
    for marker in st.session_state["markers_set"]:
        fg.add_child(marker)
        
    if st.session_state["zoom"] and not is_file_empty(land_loc):
        lat, lon, name, alt = lat_lon(st.session_state["zoom"])
        m = folium.Map(location=(lat,lon), tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', attr='Esri',zoom_start = 18)
        folium.Marker(location=[lat,lon], popup=name,icon=folium.Icon(icon = 'star',color='red')).add_to(m)
        Fullscreen().add_to(m)
        st_folium(m,feature_group_to_add=fg,key="user-map",returned_objects=[],width=900, height=575)
        st.write("Looking at:", name, " ; [Lat", lat, ",Lon:", lon, ",Alt:" ,alt, "]")

    else:
        st_folium(m,feature_group_to_add=fg,key="user-map",returned_objects=[],width=900, height=575)
        
    with st.expander("Landmarks 📍"):
        if not is_file_empty(land_loc):
            df = pd.read_csv(land_loc)
            st.write(df)

def prepare_waypoints(waypoints):
    filepath = os.path.join(parent_dir, "waypoints.csv")
    df = pd.DataFrame(waypoints)
    way_df = df[['name', 'longitude', 'latitude', 'altitud']]
    
    if os.path.exists(filepath):
        os.remove(filepath)

    way_df.to_csv(filepath, index=False) 

def draw_traj():
    filepath = os.path.join(parent_dir, "waypoints.csv")
    way_df = pd.read_csv(filepath)
    mapa2 = folium.Map(location=[40,3], tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', attr='Esri', zoom_start = 16.5, max_zoom = 20)
    Fullscreen().add_to(mapa2)

    if not way_df.empty:
        fg_waypoints = folium.FeatureGroup(name="Waypoints")
        fg_line = folium.FeatureGroup(name="Line")
        iteraciones = 0
        points = []
        for _, landmark in way_df.iterrows():
            points.append([landmark['latitude'], landmark['longitude']])
            if iteraciones == 0:         
                folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(icon = 'star',color='red')).add_to(fg_waypoints) 
            
            else:
                if iteraciones == len(way_df)-1:
                    folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(icon = 'star',color='green')).add_to(fg_waypoints) 

                else:
                    folium.Marker([landmark['latitude'], landmark['longitude']], popup=landmark['name'], icon=folium.Icon(color='orange')).add_to(fg_waypoints) 

            iteraciones += 1

        folium.PolyLine(points,color="blue",weight=1.5,opacity=1).add_to(fg_line)    
        
        mapa2.add_child(fg_waypoints)
        mapa2.add_child(fg_line)
        folium.LayerControl().add_to(mapa2)
        mapa2.fit_bounds(mapa2.get_bounds())
    st.session_state["mapa2"] = mapa2
    output = st_folium(mapa2,key="traj-map",returned_objects=[],width=850, height=700)
    
def data_uploader(name):
    filepath = os.path.join(parent_dir, name)
    try:
        df = pd.read_csv(filepath)
    except:
        df = pd.DataFrame()
        
    return df, filepath

def convert_df(data, filepath):
    df = pd.DataFrame(data)
    df.to_csv(filepath, index=False)

def increment_key():
    st.session_state["key"] += 1

def AG_lm_table(name):
    df, filepath = data_uploader(name)
    
    gd = GridOptionsBuilder.from_dataframe(df)

    gd.configure_pagination(enabled=False)
    gd.configure_default_column(editable=False, groupable=True, autoSize=True)
    gd.configure_column("name", headerCheckboxSelection=True, headerCheckboxSelectionFilteredOnly=True)
    gd.configure_selection(selection_mode='multiple')    
    gridoptions = gd.build()

    with st.form('table') as f:
        grid_table = AgGrid(df, gridOptions=gridoptions, key=st.session_state["key"],
                                editable=False,
                                allow_unsafe_jscode = True,
                                theme = 'streamlit',
                                height = 625,
                                fit_columns_on_grid_load = True)
        st.write(" *Note: Hold `Ctrl` and right-click to select multiple items*")
        sel_row = grid_table["selected_rows"]
        done = st.form_submit_button("Confirm selection 🔒", type="primary")
        
        if done:
            prepare_waypoints(sel_row)
            st.sidebar.success("Waypoints created succesfully")

def AG_table():
    df, filepath = data_uploader(st.session_state["modo"])
    z1, z2, z3, z4 = st.columns(4)
    with z1:
        _funct = st.radio('Options', options=['Display', 'Delete'])
        
    gd = GridOptionsBuilder.from_dataframe(df)

    #JavaScript injection
    cell_button_add = JsCode('''
    class BtnCellRenderer {
        init(params) {
            this.params = params;
            this.eGui = document.createElement('div');
            this.eGui.innerHTML = `
            <span>
                <style>
                .btn_add {
                    background-color: #71DC87;
                    border: 2px solid black;
                    color: #D05732;
                    text-align: center;
                    display: inline-block;
                    font-size: 12px;
                    font-weight: bold;
                    height: 2em;
                    width: 10em;
                    border-radius: 12px;
                    padding: 0px;
                }
                </style>
                <button id='click-button' 
                    class="btn_add" 
                    >&#x2193; Add</button>
            </span>
        `;
        }

        getGui() {
            return this.eGui;
        }
    };
    ''')
    
    js_add_row = JsCode('''
        function(e){
            let api = e.api;
            let rowPos = e.rowIndex + 1;
            api.applyTransaction({addIndex: rowPos, add: [{}]})
        };
        '''
        )
    
    js_delete_row = JsCode("""
        function(e) {
            let api = e.api;
            let sel = api.getSelectedRows();
            api.applyTransaction({remove: sel})
        };
        """)
    
    if _funct == 'Display':
        with z3:
            sel_mode = st.selectbox('Edit', options=['longitude', 'latitude', 'altitud'])
        
        with z2: 
            select_all = st.toggle("Select All Rows")
            
        with z4:
            entero = st.number_input("new value:", key='drop')
            drop = float(entero)
       
        gd.configure_column(field='🔧', editable=False, filter=False,
                            onCellClicked=js_add_row, cellRenderer=cell_button_add,
                            autoHeight=True, wrapText=True)
            
        gd.configure_pagination(enabled=False)
        gd.configure_default_column(editable=True, groupable=True, autoSize=True)

        if select_all:
            gd.configure_selection(selection_mode='multiple', use_checkbox=True, pre_selected_rows=df.index.tolist())  
        else:
            gd.configure_selection(selection_mode='multiple', use_checkbox=True)  

        gridoptions = gd.build()
        
        with st.form('table') as f:
            grid_table = AgGrid(df, gridOptions=gridoptions,
                                width="100%",
                                editable=True,
                                allow_unsafe_jscode = True,
                                theme = 'streamlit',
                                height = 550,
                                fit_columns_on_grid_load = True, 
                                )
            st.write(" *Note: Don't forget to hit enter ↩ on new entry.*")
            sel_row = grid_table["selected_rows"]
            done = st.form_submit_button("Confirm change(s) 🔒", type="primary")

            if done: 
                if drop and sel_mode:
                    for index in sel_row:
                        indice = index['_selectedRowNodeInfo']['nodeRowIndex']
                        if indice is not None: 
                            grid_table["data"].loc[indice, sel_mode] = drop
                
                convert_df(grid_table['data'], filepath)
                increment_key()

                    
    if _funct == 'Delete':
        gd.configure_selection('single')
        gd.configure_grid_options(onRowSelected=js_delete_row, pre_selected_rows=[])
        gridoptions = gd.build()
    
        with st.form('table') as f:
            grid_table = AgGrid(df, gridOptions=gridoptions, key='AgGrid2',
                                width="100%",
                                allow_unsafe_jscode=True,
                                enable_enterprise_modules=True,
                                fit_columns_on_grid_load=True,
                                update_mode=GridUpdateMode.SELECTION_CHANGED,
                                reload_data=True,
                                theme='streamlit',
                                height=550)
            st.write(" *Touch any row to delete*")
            done = st.form_submit_button("Confirm change(s) 🔒", type="primary")
            if done: 
                convert_df(grid_table['data'], filepath)

def menu():
    over_theme = {
        'menu_background': ' #d5d5d5 ', 
        'menu_font_size': '24px',         
        'border_width': '2px',        
        'icon_size': '24px '
    }    

    menu_data = [
    {'label':"LOAD LANDMARKS 📍"},
    {'label':"ADD WPNTS MANUALLY 🤚"},
    {'label':"CREATE 🌐"},
    {'label':"INFO  ℹ️"},
    ]

    menu_id = hc.nav_bar(menu_definition=menu_data,
                        override_theme=over_theme,
                        key='hyd') 
    return menu_id

def load_landmarks():
    c1, c2, c3 = st.columns([0.48, 0.02, 0.48])

    with st.sidebar:
        st.markdown("""
        <div style="text-align: center;">
            <h4>TOOLBAR 🛠️</h4>
        </div>
        """, unsafe_allow_html=True)  
        
        archivo = st.file_uploader('Load landmarks', accept_multiple_files=True, type=['csv', 'kml', 'kmz'])
        st.write(" ")
        delete_all = st.button("DELETE ALL ❗")

        edit_lm = st.toggle("WAYPOINTS DATA")

        if archivo:
            charge_file(archivo)
            st.sidebar.success("Landmarks added successfully")
            
        if delete_all:
            st.session_state["markers_set"].clear()
            if not is_file_empty(land_loc):
                df = pd.read_csv(land_loc)
                df.drop(df.index, inplace=True)
                df.to_csv(land_loc, index=False)
                
        #if zoom: 
            #st.session_state["zoom"] = zoom_select

    with c1: 
        st.markdown("""
        <div style="text-align: center;">
            <h4>LANDMARK MAP 📍</h4>
        </div>
        """, unsafe_allow_html=True)  
        draw_map()
    
    if edit_lm:
        with c3: 
            st.markdown("""
            <div style="text-align: center;">
                <h4>LANDMARK DATA</h4>
            </div>
            """, unsafe_allow_html=True)  
            
            st.session_state["modo"] = "landmarks.csv"
            st.fragment(AG_table, run_every=2.0)()
            #AG_table("landmarks.csv")
    
def info():
    st.info("DEVELOPING")

def clear_waypoints():
    filepath = os.path.join(parent_dir, "waypoints.csv")

    try: 
        df = pd.read_csv(filepath)
        columns = df.columns
        empty_df = pd.DataFrame(columns=columns)
        empty_df.to_csv(filepath, index=False)
        increment_key()
        
    except: 
        st.sidebar.error("Error: can´t delete wpnts.")
    
@st.dialog("SUCCESS")
def traj_created(name):
    st.info(f"Trajectory {name} created correctyl")
    st.info(f"Dont forget to save {name} in trajectories folder")
    time.sleep(2)

@st.dialog("FAILURE")
def traj_not_created(name):
    st.error(f"Trajectory {name} already exists, please enter a non existing name")
    
def create_trajectory():
    c1, c2, c3 = st.columns([0.48, 0.02,0.48])
    
    with st.sidebar: 
        st.markdown("""
        <div style="text-align: center;">
            <h4>TOOLBAR 🛠️</h4>
        </div>
        """, unsafe_allow_html=True)  
        a1, a2 = st.columns(2, vertical_alignment="bottom")
        with a1: 
            st.session_state["name"] = st.text_input("name:")
            name = st.session_state["name"]
            
        with a2:
            create = st.button("CREATE❗")
            if create and not st.session_state["name"]:
                st.sidebar.error("Please enter a name before adding")
        
        st.markdown("""
        <div style="text-align: center;">
            <h3 style="font-size: 16px;">LANDMARKS</h3>
        </div>
        """, unsafe_allow_html=True) 
        b1, b2 = st.columns(2, vertical_alignment="bottom")
        with b1:
            add_lm = st.toggle("Add") 
            
        with b2: 
            info_lm = st.button("info", key="edit1")
            
        st.write(" ")
        st.markdown("""
        <div style="text-align: center;">
            <h3 style="font-size: 16px;">WAYPOINTS</h3>
        </div>
        """, unsafe_allow_html=True) 
        d1, d2 = st.columns(2, vertical_alignment="bottom")
        with d1:
            manual = st.toggle("show")
            
        with d2: 
            auto_wyd = st.button("🤖", key="edit2")
            
        st.write(" ")
        st.markdown("""
        <div style="text-align: center;">
            <h3 style="font-size: 16px;">TRAJECTORY</h3>
        </div>
        """, unsafe_allow_html=True) 
        show_traj = st.toggle("show 🗺️")
    
        clear = st.button("CLEAR WPNTS ❌ ")
        
        if st.session_state["markers_set"] and add_lm:
            with c3: 
                st.markdown(f"""
                <div style="text-align: center;">
                <h4>LANDMARKS {name}</h4>
                </div>
                """, unsafe_allow_html=True)  
                st.write("---")
                AG_lm_table("landmarks.csv")
        
        if auto_wyd or info_lm:
            st.sidebar.info("UNDER DEVELOPMENT")
        
        df = data_uploader("waypoints.csv")
        if not df[0].empty:
            if not add_lm and manual:
                with c3: 
                    st.markdown(f"""
                    <div style="text-align: center;">
                        <h3>WAYPOINTS {name} 🛰️</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.session_state["modo"] = "waypoints.csv"
                    st.fragment(AG_table, run_every=2.0)()
                    #AG_table("waypoints.csv")
                    
            if clear: 
                clear_waypoints()
            
            if show_traj: 
                with c1:
                    st.markdown(f"""
                    <div style="text-align: center;">
                        <h3>TRAJECTORY {name} 🚁</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    st.fragment(draw_traj, run_every=1.0)()
                    
            if create and not st.session_state["name"]:
                st.sidebar.error("Name not set")
            else:
                carpeta = os.path.join(parent_dir, "Trayectorias")
                ruta_carpeta = Path(carpeta)
                directorios = [item.name for item in ruta_carpeta.iterdir() if item.is_dir()]
                if create and st.session_state["name"] and st.session_state["name"] not in directorios:
                    prepare_data_route(st.session_state["name"])
                    way_path = os.path.join(parent_dir, "waypoints.csv")
                    new_path = os.path.join(parent_dir, "Trayectorias", st.session_state["name"], "waypoints.csv")
                    way_df = pd.read_csv(way_path)
                    way_df.to_csv(new_path, index=False)
                    plantilla = os.path.join(parent_dir, "Trayectorias", st.session_state["name"], st.session_state["name"] + ".xlsx")
                    wb = load_workbook(plantilla)
                    ws = wb.active
                    cor1 = devolver_lat_lon(way_df, 0)
                    ws['D5'] = ','.join(cor1)
                    cor2 = devolver_lat_lon(way_df, 1)
                    ws['E5'] = ','.join(cor2)
                    wb.save(plantilla)
                    mapa_path = os.path.join(parent_dir, "Trayectorias", st.session_state["name"], st.session_state["name"]+".html")
                    st.session_state["mapa2"].save(mapa_path)
                    wb.close()
                    st.session_state["waypoints"] = []
                    st.session_state["name"] = None
                    st.session_state["route_name"] = None
                    st.session_state["new"] = False
                    traj_created(name)
                else: 
                    if st.session_state["name"] in directorios and create:
                        traj_not_created(name)
                
        else:
            if add_lm:
                st.sidebar.info("Desactiva el toggle del landmarks para empezar a trabajar en la trayectoria")

            if manual: 
                st.sidebar.info("No landmarks selected")
                
            if df[0].empty and clear:
                st.sidebar.info("No waypoints yet")
                
@st.cache_data
def base_manual_map(coord=[40,3]):
    m = folium.Map(location=coord, tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', attr='Esri', zoom_start = 16.5, max_zoom = 20)
    draw = Draw(
        draw_options={
            'polyline': {
                'shapeOptions': {
                    'color': '#00FF00', 
                    'weight': 6        
                }
            },
            'polygon': True,
            'circle': False,
            'rectangle': False,
            'marker': False,
            'circlemarker': False,
        },
        edit_options={
            'poly': {
                'shapeOptions': {
                    'color': '#00FFFF',
                    'weight': 6
                }
            }
        }
    )
    draw.add_to(m) 
    Fullscreen().add_to(m)

    return m

#@st.cache_data
def refresh_ouput(output):
    st.session_state["manual_set"].clear()
    st.session_state["df_manual"].clear()
    if output:
        number_name = 0
        if "all_drawings" in output:
            all_drawings = output["all_drawings"]
            if isinstance(all_drawings, list):
                for drawing in all_drawings:
                    counter = 0
                    draw_last = all_drawings[-1]
                    draw_first = all_drawings[0]
                    if drawing.get("geometry", {}).get("type") == "LineString":
                        coordinates = drawing.get("geometry", {}).get("coordinates", [])
                        for coord in coordinates:
                            if counter == 0 and drawing['geometry']['coordinates'][0] == draw_first['geometry']['coordinates'][0]:
                                color = "red"
                            elif counter == len(coordinates)-1 and drawing['geometry']['coordinates'][-1] == draw_last['geometry']['coordinates'][-1]:
                                color = "green"
                            else:
                                color = "orange"
                            data = {
                                "latitude": coord[1],
                                "longitude": coord[0],
                                "name": f"waypoint {number_name}",
                                "altitud": 100
                            }
                            st.session_state["manual_set"].append(create_marker(data, color))

                            st.session_state["df_manual"].append({
                            "name": f"waypoint {number_name}",
                            "latitude": coord[1],
                            "longitude": coord[0],
                            "altitud": 100
                            })
                            
                            number_name += 1
                            counter += 1
          
@st.fragment(run_every=1.75)
def manual_map():
    m = base_manual_map([st.session_state["data_map"][0], st.session_state["data_map"][1]])
    
    #Si devuelvo solo polyline sale vacío aunque dibuje, no se por qué
    st.session_state["output"] = st_folium(m,key="user-map-manual",returned_objects=["all_drawings"],width=900, height=750)
    refresh_ouput(st.session_state["output"])

def visual_map():
    coordinates = []
    m = base_map([st.session_state["data_map"][0], st.session_state["data_map"][1]], 16.5)

    fg = folium.FeatureGroup(name="Markers")

    for marker in st.session_state["manual_set"]:
        fg.add_child(marker)
        if hasattr(marker, 'location'):
            coordinates.append(marker.location)

    # Crear y añadir la polilínea si hay suficientes puntos
    if len(coordinates) > 1:
        folium.PolyLine(
            locations=coordinates,  
            color='blue',          
            weight=2,               
            opacity=0.8            
        ).add_to(fg)

    #m.fit_bounds(m.get_bounds())

    st_folium(m,key="user-map-visual",feature_group_to_add=fg, width=900, height=750)

@st.dialog("SUCCESS")
def manual_success():
    st.success("Waypoints loaded correctly")
    st.info("CREATE => WAYPOINTS => MANUAL")

def split_data(df, chunk_size):
    return np.array_split(df, np.ceil(len(df) / chunk_size))

def simple_process(url, coords):
    all_results = []
    #Dividir las coordenadas en chunks de 100
    for i, coords_chunk in tqdm(coords.groupby(coords.index // 100)):
        time.sleep(1.1)
        coords_str = coords_chunk.to_string(col_space=1, index=False, header=False)
        coords_str = ",".join(coords_str.replace('\n', '|').split())

        req_data = {"locations": coords_str, "interpolation": "bilinear"}
        r = requests.post(url, data=req_data)

        if r.status_code == 200: 
            print("CHUNK OK")
        else:
            print("status:")
            print(r.status_code)

        all_results += r.json()['results']

    return all_results

#@st.cache_data // rentaría ? 
def ASL(df):
    coords = df[['latitude', 'longitude']].copy()

    # URL de la API de Open Topo Data
    url = "https://api.opentopodata.org/v1/eudem25m"
    #all_results = process_requests(coords, url, max_calls_per_block=5, max_threads=2)
    all_results = []
    if len(coords) > 499:
        chunk_size = 499
        chunks = split_data(coords, chunk_size)
        for i, chunk in enumerate(chunks):
            #time.sleep(5)
            all_results.extend(simple_process(url, chunk))

    else:
        all_results = simple_process(url, coords)

    elevations = [res['elevation'] for res in all_results]
    elevations = [round(elevation, 6) for elevation in elevations]

    df['ASL'] = elevations

    return df

def manual():
    c1,c0, c2 = st.columns([0.49, 0.02, 0.49])

    with st.sidebar:
        st.markdown("""
        <div style="text-align: center;">
            <h4>TOOLBAR 🛠️</h4>
        </div>
        """, unsafe_allow_html=True)
        way_path = os.path.join(parent_dir, "waypoints.csv")
        
        man = st.toggle("MANUAL")
        auto = st.toggle("AUTO FILL")
        load_wypnts = st.button("LOAD WYPNTS")
        st.write("")
        planta = st.text_input("SEARCH PLANT")

        if planta:
            planta_lower = planta.lower()
            for index, row in st.session_state["plantas"].iterrows():
                if row["name"].lower() == planta_lower:
                    st.session_state["data_map"] = [row["latitude"], row["longitude"]]
                    break
                else:
                    if row["name"].lower() != planta_lower and index == len(st.session_state["plantas"]) - 1:
                        st.sidebar.error(f"No plant {planta}")
                 
        if load_wypnts:
            way_path = os.path.join(parent_dir, "waypoints.csv")
            df = pd.DataFrame(st.session_state["df_manual"])
            
            df_final = ASL(df)
            
            if os.path.exists(way_path):
                os.remove(way_path)

            df_final.to_csv(way_path, index=False)
            st.session_state["df_manual"].clear()
            manual_success()

        if man:
            st.session_state["modo_man"] = 1
        elif auto:
            st.info("DEVELOPING")
            st.session_state["modo_man"] = 2
        else:
            st.info("NO MODE SELECTED")
            st.session_state["modo_man"] = 0
            
    with c1:
        st.markdown("""
        <div style="text-align: center;">
            <h4>INTERACTIVE MAP 🎨</h4>
        </div>
        """, unsafe_allow_html=True)   
        st.write(" ")
        if man:
            manual_map()
        else: 
            st.info("NO MODE SELECTED // Only manual mode available at the moment")
        
    with c2:
        st.markdown("""
            <div style="text-align: center;">
                <h4> OUTPUT MAP 🗺️</h4>
            </div>
            """, unsafe_allow_html=True)   
        st.write(" ")
        if man:
            st.fragment(visual_map, run_every=2.0)()

def main():
    prepare_data()
    selected = menu()
    
    if selected == "LOAD LANDMARKS 📍":
        load_landmarks()
    
    if selected == "CREATE 🌐":
        create_trajectory()
    
    if selected == "ADD WPNTS MANUALLY 🤚":
        manual()
        
    if selected == "INFO  ℹ️":
        info()
        
main()